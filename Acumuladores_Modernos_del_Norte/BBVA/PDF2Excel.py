# This script converts data from a PDF file to an Excel file based on specific keywords and patterns using Regex. Extracting relevant 
# information such as dates, financial movements, references, and beneficiaries. The extracted data is # then saved to an Excel file.

# External libraries
import pdfplumber
import openpyxl
import re

# PDF Text extraction Function
def pdf_text_extraction(pdf_path):
    # Open the PDF file and extract all text from its pages
    with pdfplumber.open(pdf_path) as pdf:
        text = ""
        for page in pdf.pages:
            text += page.extract_text()
    return text

# Line rgister function
def information_extraction(text, key_words):
    # Split the text into lines for easier processing
    Lines = text.split('\n')
    results = []

    # Iterate through lines and analyze each one for the presence of key words    
    for i in range(len(Lines) - 1):
        Line = Lines[i]
        next_line = Lines[i + 1]
        for word in key_words:
            if word in Line:

                # Extract the date (assumes it is the first element in the line)
                date_str = Line.split()[0]
                Date = int(date_str.split('/')[0])
                
                # Extract financial movement using regular expressions
                number = re.findall(r'\d{1,3}(?:,\d{3})*(?:\.\d{2})?', Line)
                if len(number) >= 4:
                    movement = number[3]
                elif len(number) >= 1:
                    movement = number[0]
                    
                # Extract the reference, isolating it between the date and the movement
                reference_1 = Line.split(date_str, 1)[1].split(movement, 1)[0].strip()
                # Clean the reference further by removing specific patterns (e.g., "XX/MMM")
                reference_2 = re.sub(r'\d{2}/[A-Z]{3}', '', reference_1).strip()
                
                # Save the result, including date, movement, reference, and beneficiary (next line)
                result = (Date, movement, reference_2, next_line)
                results.append(result)
                
                break
    
    return results

# Excel File saving function
def Excel_saving(results, excel_path):
    # Create a new Excel workbook and activate the default sheet
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Resultados"
    
    # Set column titles for the results
    sheet.cell(row=1, column=1).value = "Fecha"
    sheet.cell(row=1, column=2).value = "Movimiento"
    sheet.cell(row=1, column=3).value = "Referencia"
    sheet.cell(row=1, column=4).value = "Beneficiario"
    
    # Populate the Excel sheet with extracted data
    for idx, (fecha, movimiento, referencia, Beneficiario) in enumerate(results, start=2):
        sheet.cell(row=idx, column=1).value = fecha
        sheet.cell(row=idx, column=2).value = movimiento
        sheet.cell(row=idx, column=3).value = referencia
        sheet.cell(row=idx, column=4).value = Beneficiario
    
    # Save the Excel file to the specified path
    book.save(excel_path)
    print(f"PDF convertido a Excel exitosamente: {excel_path}")

# Main PDF to Excel function
def PDF2Excel(pdf_path, key_words, excel_path):
    
    # Extract text from the PDF
    pdf_text = pdf_text_extraction(pdf_path)

    # Search for key words and extract relevant information from the text
    Data = information_extraction(pdf_text, key_words)

    # Save the extracted data to an Excel file
    Excel_saving(Data, excel_path)