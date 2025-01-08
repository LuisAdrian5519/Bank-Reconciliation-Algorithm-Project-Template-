# This file is used to extract the data from an Excel database file: Date, Beneficiary, Reference, Income, and Outcome.
# And return it to the main algorithm
 
# External libraries
import openpyxl

def Value_extraction(key_words_Ingresos, key_words_Egresos, Excel_path, Columna_para_fechas, Columna_para_movimientos, Columna_para_referencias, Columna_para_beneficiario, HEADER):
   
    # Variables declaration
    Ingresos = []                   # Income transactions
    Egresos = []                    # Outcome transactions
   
    Fechas_Ingresos = []            # Income transaction dates
    Fechas_Egresos = []             # Outcome transaction dates
   
    Beneficiarios_Ingresos = []     # Income transaction beneficiaries
    Beneficiarios_Egresos = []      # Outcome transaction beneficiaries

    Referencias_Ingresos = []       # Income transaction references
    Referencias_Egresos = []        # Outcome transaction references

    # Excel Preprocessing - Opening the Excel file and accessing the active sheet
    Excel = openpyxl.load_workbook(Excel_path)
    Dataframe = Excel.active

    # Excel Processing - Iterating through rows to extract and classify data
    for row in Dataframe.iter_rows(min_row=HEADER, max_row=Dataframe.max_row):

        # Extract and preprocess data from the specified columns
        fecha = int(str(row[Columna_para_fechas].value).strip())                                # Extract the date and convert it to an integer
        referencia = row[Columna_para_referencias].value                                        # Extract the reference string
        beneficiario = row[Columna_para_beneficiario].value                                     # Extract the beneficiary string

        movimiento = float(str(row[Columna_para_movimientos].value).strip().replace(',', ''))   # Extract and convert the movement amount to a float

        # Classify as income if any keyword from key_words_Ingresos is in the reference
        if any(ingreso in referencia for ingreso in key_words_Ingresos):
            Ingresos.append(movimiento)
            Fechas_Ingresos.append(fecha)
            Beneficiarios_Ingresos.append(beneficiario)
            Referencias_Ingresos.append(referencia)

        # Classify as outcome if any keyword from key_words_Egresos is in the reference
        elif any(egreso in referencia for egreso in key_words_Egresos):
            Egresos.append(movimiento)
            Fechas_Egresos.append(fecha)
            Beneficiarios_Egresos.append(beneficiario)
            Referencias_Egresos.append(referencia)
    
    return Ingresos, Egresos, Fechas_Ingresos, Fechas_Egresos, Referencias_Ingresos, Referencias_Egresos, Beneficiarios_Ingresos, Beneficiarios_Egresos