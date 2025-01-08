# This file is used to extract the data from an Excel database file: Date, Beneficiary, Reference, Income, and Outcome.
# And return it to the main algorithm

# External libraries
import openpyxl
from datetime import datetime

def Value_extraction():

    # INSERT DATA HERE  :D
    # Replace the following variables with your specific inputs:   
    Nombre_del_archivo_Excel = "My_auxiliary_register.xls"       # File name
    Columna_para_fechas = 1                                      # Excel Column for Dates data
    Columna_para_beneficiario = 5                                # Excel Column for Beneficiary data
    Columna_para_referencias = 6                                 # Excel Column for References data
    Columna_para_ingresos = 9                                    # Excel Column for Incomes data
    Columna_para_egresos = 8                                     # Excel Column for Outcomes data
   
    HEADER = 19                                                  # Rows over the table unnecessary for analysis

    # Variables declaration
    Ingresos = []                        # Income transactions
    Ingresos_aux = []                    # Auxiliary list for income transactions
   
    Egresos = []                         # Outcome transactions
   
    Fechas = []                          # Transaction dates
    Fechas_Ingresos = []                 # Income transaction dates
    Fechas_Egresos = []                  # Outcome transaction dates
   
    Beneficiarios_General = []           # Beneficiaries of the transactions
    Beneficiarios_Ingresos = []          # Income transaction beneficiaries
    Beneficiarios_Egresos = []           # Outcome transaction beneficiaries
   
    Referencias_General_Auxiliar = []    # References of the transactions
    Referencias_Ingresos = []            # Income transaction references
    Referencias_Egresos = []             # Outcome transaction references

    # Excel Preprocessing - Open the specified Excel file and load the active sheet
    Excel = openpyxl.load_workbook(Nombre_del_archivo_Excel)
    Dataframe = Excel.active

    # Excel Processing - Iterate over each row to extract data from specified columns
    for row in Dataframe.iter_rows(HEADER, Dataframe.max_row):
        # This loop reads data row by row, extracting dates, incomes, expenses, beneficiaries, 
        # # and references into corresponding lists for further processing.
        Fechas.append(row[Columna_para_fechas].value)
        Beneficiarios_General.append(row[Columna_para_beneficiario].value)
        Referencias_General_Auxiliar.append(row[Columna_para_referencias].value)
    
        # Check if the income value is not None and is a valid number (int or float)
        # Append it to the auxiliary incomes list if valid
        valor_ingreso = row[Columna_para_ingresos].value
        if valor_ingreso is not None and isinstance(valor_ingreso, (int, float)):
            Ingresos_aux.append(valor_ingreso)
        
        Ingreso = row[Columna_para_ingresos].value
        Egreso = row[Columna_para_egresos].value

        # Extract and validate the income and outcome value, defaulting to 0 if not a valid number
        if isinstance(Ingreso, (int, float)):
            Ingresos.append(Ingreso)
        else:
            Ingresos.append(0)
        
        if isinstance(Egreso, (int, float)):
            Egresos.append(Egreso)
        else:
            Egresos.append(0)    
  
    # Process dates to extract the day from each date
    Dias = [Fecha.day for Fecha in Fechas if Fecha is not None and isinstance(Fecha, datetime)]

    Ingresos = [valor for valor in Ingresos if valor != 0]
    Egresos = [valor for valor in Egresos if valor != 0]

    # Excel Postprocessing - Separation into income and outcome transactions
    for i in range(len(Ingresos_aux)):
        # This loop checks each transaction in the auxiliary incomes list.
        # If the value is zero, it is classified as an outcome; otherwise, it is an income.
        if Ingresos_aux[i] == 0:
            Fechas_Egresos.append(Dias[i])
            Referencias_Egresos.append(Referencias_General_Auxiliar[i])
            Beneficiarios_Egresos.append(Beneficiarios_General[i])
           
        else:
            Fechas_Ingresos.append(Dias[i])
            Referencias_Ingresos.append(Referencias_General_Auxiliar[i])
            Beneficiarios_Ingresos.append(Beneficiarios_General[i])
    
    return Ingresos, Egresos, Fechas_Ingresos, Fechas_Egresos, Referencias_Ingresos, Referencias_Egresos, Beneficiarios_Ingresos, Beneficiarios_Egresos