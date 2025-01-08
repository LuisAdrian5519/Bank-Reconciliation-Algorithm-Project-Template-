# This file is used to extract the data from an Excel database file: Date, Beneficiary, Reference, Income, and Outcome.
# And return it to the main algorithm

# External libraries
import openpyxl

def Value_extraction():

    # INSERT DATA HERE  :D
    # Replace the following variables with your specific inputs:   
    Nombre_del_archivo_Excel = "My_auxiliary_register.xls"       # File name
    Columna_para_fechas = 0                                      # Excel Column for Dates data
    Columna_para_beneficiario = 5                                # Excel Column for Beneficiary data
    Columna_para_referencias = 7                                 # Excel Column for References data
    Columna_para_ingresos = 8                                    # Excel Column for Incomes data
    Columna_para_egresos = 9                                     # Excel Column for Outcomes data
   
    HEADER = 1                                                   # Rows over the table unnecessary for analysis
   
    # Variables declaration
    Ingresos = []                        # Income transactions
    Ingresos_aux = []                    # Auxiliary list for income transactions
   
    Egresos = []                         # Outcome transactions
   
    Fechas = []                          # Transaction dates
    Fechas_Ingresos = []                 # Income transaction dates
    Fechas_Egresos = []                  # Outcome transaction dates
   
    Beneficiarios_General = []           # Beneficiaries of the transactions
    Beneficiarios_Ingresos = []          # Income transaction beneficiaries
    Beneficiarios_Egresos = []           # Outcome transaction beneficiaries
   
    Referencias_General_Auxiliar = []    # References of the transactions
    Referencias_Ingresos = []            # Income transaction references
    Referencias_Egresos = []             # Outcome transaction references

    # Excel Preprocessing - Open the specified Excel file and load the active sheet
    Excel = openpyxl.load_workbook(Nombre_del_archivo_Excel)
    Dataframe = Excel.active

    # Excel Processing - Iterate over each row to extract data from specified columns
    for row in Dataframe.iter_rows(HEADER, Dataframe.max_row):
        # This loop reads data row by row, extracting dates, incomes, expenses, beneficiaries, 
        # and references into corresponding lists for further processing.
        Fechas.append(row[Columna_para_fechas].value)
        Ingresos.append(row[Columna_para_ingresos].value)
        Ingresos_aux.append(row[Columna_para_ingresos].value)
        Egresos.append(row[Columna_para_egresos].value)
        Beneficiarios_General.append(row[Columna_para_beneficiario].value)
        Referencias_General_Auxiliar.append(row[Columna_para_referencias].value)
       
    # Process dates to extract the day from each date
    Dias = [Fecha.day for Fecha in Fechas if Fecha is not None]
   
    Ingresos = [valor for valor in Ingresos if valor != 0]
    Egresos = [valor for valor in Egresos if valor != 0]
   
    # Excel Postprocessing - Separation into income and outcome transactions
    for i in range(len(Ingresos_aux)):
        # This loop checks each transaction in the auxiliary incomes list.
        # If the value is zero, it is classified as an outcome; otherwise, it is an income.
        if Ingresos_aux[i] == 0:
            Fechas_Egresos.append(Dias[i])
            Referencias_Egresos.append(Referencias_General_Auxiliar[i])
            Beneficiarios_Egresos.append(Beneficiarios_General[i])
           
        else:
            Fechas_Ingresos.append(Dias[i])
            Referencias_Ingresos.append(Referencias_General_Auxiliar[i])
            Beneficiarios_Ingresos.append(Beneficiarios_General[i])
    
    return Ingresos, Egresos, Fechas_Ingresos, Fechas_Egresos, Referencias_Ingresos, Referencias_Egresos, Beneficiarios_Ingresos, Beneficiarios_Egresos